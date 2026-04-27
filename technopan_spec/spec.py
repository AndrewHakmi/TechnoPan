from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .dxf import PanelItem


@dataclass(frozen=True)
class PanelRow:
    idx: int
    supply_no: str
    panel_type: str
    tag_prefix: str | None          # буква маркировки, напр. "п"
    tag_number: int | None          # цифра маркировки, напр. 612
    ral_out: str | None
    metal_out_mm: float | None
    profile_out: str | None
    ral_in: str | None
    metal_in_mm: float | None
    profile_in: str | None
    length_mm: float
    width_mm: float
    thickness_mm: float
    qty: float
    area_m2_total: float
    coating_out: str | None
    coating_in: str | None


def _group_key(i: PanelItem) -> tuple:
    return (
        i.panel_type,
        i.tag_prefix,
        i.tag_number,
        i.ral_out,
        i.metal_out_mm,
        i.profile_out,
        i.coating_out,
        i.ral_in,
        i.metal_in_mm,
        i.profile_in,
        i.coating_in,
        i.length_mm,
        i.width_mm,
        i.thickness_mm,
    )


def build_panel_rows(items: list[PanelItem]) -> list[PanelRow]:
    grouped: dict[tuple, dict[str, float]] = {}
    for i in items:
        key = _group_key(i)
        g = grouped.setdefault(key, {"qty": 0.0, "area": 0.0})
        g["qty"] += float(i.qty)
        # Round area per panel to 3 decimals first to match manual Excel behavior
        panel_area = round(float(i.length_mm) * float(i.width_mm) / 1_000_000.0, 3)
        g["area"] += panel_area * float(i.qty)

    rows: list[PanelRow] = []
    for n, (key, agg) in enumerate(sorted(grouped.items(), key=lambda kv: kv[0]), start=1):
        (
            panel_type,
            tag_prefix,
            tag_number,
            ral_out,
            metal_out_mm,
            profile_out,
            coating_out,
            ral_in,
            metal_in_mm,
            profile_in,
            coating_in,
            length_mm,
            width_mm,
            thickness_mm,
        ) = key

        rows.append(
            PanelRow(
                idx=n,
                supply_no="ПК-",
                panel_type=str(panel_type),
                tag_prefix=tag_prefix,
                tag_number=tag_number,
                ral_out=ral_out,
                metal_out_mm=metal_out_mm,
                profile_out=profile_out,
                ral_in=ral_in,
                metal_in_mm=metal_in_mm,
                profile_in=profile_in,
                length_mm=float(length_mm),
                width_mm=float(width_mm),
                thickness_mm=float(thickness_mm),
                qty=round(float(agg["qty"]), 3),
                area_m2_total=round(float(agg["area"]), 3),
                coating_out=coating_out,
                coating_in=coating_in,
            )
        )
    return rows


EXPORT_COLUMNS = [
    ("idx", "№ п.п.", 8),
    ("supply_no", "№ поставки", 12),
    ("panel_type", "Тип панели", 20),
    ("tag_prefix", "Маркировка (буква)", 18),
    ("tag_number", "Маркировка (номер)", 18),
    ("ral_out", "RAL наруж", 18),
    ("metal_out_mm", "Толщина металла наруж, мм", 18),
    ("profile_out", "Профилирование наруж", 18),
    ("ral_in", "RAL внутр", 18),
    ("metal_in_mm", "Толщина металла внутр, мм", 18),
    ("profile_in", "Профилирование внутр", 18),
    ("length_mm", "Длина, мм", 18),
    ("width_mm", "Ширина, мм", 18),
    ("thickness_mm", "Толщина, мм", 18),
    ("qty", "Кол-во, шт.", 18),
    ("area_m2_total", "Площадь, м2 общая", 18),
    ("coating_out", "Покрытие наруж", 18),
    ("coating_in", "Покрытие внутр", 18),
]

def write_spec_xlsx(path: Path, rows: list[PanelRow], title: str, active_columns: list[str] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Панели"

    if active_columns is None:
        active_columns = [c[0] for c in EXPORT_COLUMNS]

    # Filter columns
    cols_to_write = [c for c in EXPORT_COLUMNS if c[0] in active_columns]
    headers = [c[1] for c in cols_to_write]
    col_ids = [c[0] for c in cols_to_write]

    if not headers:
        wb.save(path)
        return

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, row in enumerate(rows, start=4):
        for c, col_id in enumerate(col_ids, start=1):
            val = getattr(row, col_id, None)
            ws.cell(row=r, column=c, value=val)

    total_row = 4 + len(rows)
    ws.cell(row=total_row, column=1, value="ИТОГО")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # Place totals in the correct columns
    if "qty" in col_ids:
        c_idx = col_ids.index("qty") + 1
        total_qty = sum(r.qty for r in rows)
        ws.cell(row=total_row, column=c_idx, value=round(total_qty, 3))
        ws.cell(row=total_row, column=c_idx).font = Font(bold=True)
        
    if "area_m2_total" in col_ids:
        c_idx = col_ids.index("area_m2_total") + 1
        total_area = sum(r.area_m2_total for r in rows)
        ws.cell(row=total_row, column=c_idx, value=round(total_area, 3))
        ws.cell(row=total_row, column=c_idx).font = Font(bold=True)

    ws.freeze_panes = "A4"

    from openpyxl.utils import get_column_letter
    for c, col_def in enumerate(cols_to_write, start=1):
        width = col_def[2]
        ws.column_dimensions[get_column_letter(c)].width = width

    wb.save(path)

